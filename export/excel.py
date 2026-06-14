"""
excel.py
Step 3 — Excel export
Takes identity + phones and builds a clean Excel file with confidence scores
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ── Colors ────────────────────────────────────────────────────────────────────
COLOR_HEADER   = "1F3864"  # Dark blue
COLOR_HIGH     = "C6EFCE"  # Green  — confidence >= 70%
COLOR_MEDIUM   = "FFEB9C"  # Yellow — confidence 40-69%
COLOR_LOW      = "FFC7CE"  # Red    — confidence < 40%
COLOR_ALT_ROW  = "F2F2F2"  # Light grey for alternating rows

MAX_PHONES = 3  # Max phone columns per company


# ── Helpers ───────────────────────────────────────────────────────────────────

def confidence_color(score: int) -> str:
    if score >= 70:
        return COLOR_HIGH
    elif score >= 40:
        return COLOR_MEDIUM
    else:
        return COLOR_LOW


def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def style_data_cell(cell, value, alt_row=False, bg_color=None):
    cell.value = value
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = make_border()
    cell.font = Font(size=10)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    elif alt_row:
        cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)


# ── Main Export Function ──────────────────────────────────────────────────────

def export_to_excel(leads: list, output_path: str = None) -> str:
    """
    Export a list of enriched leads to Excel.

    Each lead is a dict with:
        identity: dict from identity.py
        phones:   list of dicts from dm_hunter.py

    Args:
        leads:       List of enriched lead dicts
        output_path: Optional custom output path

    Returns:
        Path of the created Excel file
    """

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"leads_export_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads B2B"

    # ── Build headers ─────────────────────────────────────────────────────────
    base_headers = [
        "Nom", "SIREN", "SIRET", "NAF", "Gérant/CEO", "Adresse", "Site Web"
    ]
    phone_headers = []
    for i in range(1, MAX_PHONES + 1):
        phone_headers += [
            f"Téléphone {i}",
            f"Type {i}",
            f"Confiance {i}",
            f"Source {i}",
        ]

    all_headers = base_headers + phone_headers
    ws.row_dimensions[1].height = 35

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        style_header_cell(cell, header)

    # ── Write data rows ───────────────────────────────────────────────────────
    for row_idx, lead in enumerate(leads, 2):
        identity = lead.get("identity", {})
        phones   = lead.get("phones", [])
        alt      = row_idx % 2 == 0

        ws.row_dimensions[row_idx].height = 20

        # Base identity columns
        base_values = [
            identity.get("nom", "N/A"),
            identity.get("siren", "N/A"),
            identity.get("siret", "N/A"),
            identity.get("naf", "N/A"),
            identity.get("ceo", "N/A"),
            identity.get("address", "N/A"),
            identity.get("website", "N/A"),
        ]

        for col_idx, val in enumerate(base_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            style_data_cell(cell, val, alt_row=alt)

        # Phone columns
        for phone_idx in range(MAX_PHONES):
            base_col = len(base_headers) + (phone_idx * 4) + 1

            if phone_idx < len(phones):
                p = phones[phone_idx]
                score = p.get("confidence", 0)
                color = confidence_color(score)

                values = [
                    p.get("phone", ""),
                    p.get("type", ""),
                    f"{score}%",
                    p.get("sources", ""),
                ]
            else:
                color = None
                values = ["", "", "", ""]

            for offset, val in enumerate(values):
                cell = ws.cell(row=row_idx, column=base_col + offset)
                style_data_cell(cell, val, alt_row=alt, bg_color=color if offset == 2 else None)

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [30, 12, 16, 8, 25, 40, 30] + [18, 10, 12, 25] * MAX_PHONES

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Résumé")
    ws2["A1"] = "Export B2B Lead Scraper"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Date export"
    ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A4"] = "Total leads"
    ws2["B4"] = len(leads)
    ws2["A5"] = "Leads avec téléphone"
    ws2["B5"] = sum(1 for l in leads if l.get("phones"))
    ws2["A6"] = "Taux couverture"
    total = len(leads)
    covered = sum(1 for l in leads if l.get("phones"))
    ws2["B6"] = f"{round(covered/total*100)}%" if total > 0 else "0%"

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(output_path)
    logger.info(f"✅ Excel saved: {output_path}")
    return output_path


# ── Quick Test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    test_leads = [
        {
            "identity": {
                "nom": "KANDBAZ",
                "siren": "497933408",
                "siret": "49793340800054",
                "naf": "74.90B",
                "ceo": "N/A",
                "address": "1 RUE DE STOCKHOLM 75008 PARIS",
                "website": "https://www.kandbaz.com",
            },
            "phones": [
                {
                    "phone": "01 44 70 70 70",
                    "type": "Fixe",
                    "confidence": 55,
                    "sources": "Website (/contact)",
                }
            ],
        },
        {
            "identity": {
                "nom": "BNP PARIBAS",
                "siren": "662042449",
                "siret": "66204244900014",
                "naf": "64.19Z",
                "ceo": "ASCHENBROICH",
                "address": "16 BOULEVARD DES ITALIENS 75009 PARIS",
                "website": "N/A",
            },
            "phones": [],
        },
    ]

    path = export_to_excel(test_leads, "test_export.xlsx")
    print(f"\n✅ File created: {path}")
    print("Open it in Excel to check the output.")